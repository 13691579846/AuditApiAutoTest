"""
------------------------------------
@Time : 2019/6/13 9:01
@Auth : linux超
@File : ParseExcel.py
@IDE  : PyCharm
@Motto: Real warriors,dare to face the bleak warning,dare to face the incisive error!
------------------------------------
"""
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles.colors import (BLACK, RED, GREEN)
from collections import namedtuple

from config.config import DATA_PATH


class ParseExcel(object):
    def __init__(self, filename):
        self.filename = filename
        self.wb = load_workbook(self.filename)
        self.font = Font(color=None)
        self.RGBDict = {'red': RED, 'green': GREEN, 'black': BLACK}

    def excel_max_row(self, sheet_name):
        """get max row of sheet"""
        return self.wb[sheet_name].max_row

    def excel_min_row(self, sheet_name):
        """get min row of sheet"""
        return self.wb[sheet_name].min_row

    def excel_max_col(self, sheet_name):
        """get max column of sheet"""
        return self.wb[sheet_name].max_column

    def excel_min_col(self, sheet_name):
        """get min column of sheet"""
        return self.wb[sheet_name].min_column

    def head(self, sheet_name):
        """get head of excel"""
        min_row = self.excel_min_row(sheet_name)
        head = tuple(self.wb[sheet_name].iter_rows(max_row=min_row, values_only=True))[0]
        return head

    def get_all_values(self, sheet_name):
        """get name tuple of all values for excel"""
        case_list = []
        head = self.head(sheet_name)
        case = namedtuple('case', head)
        all_values = self.wb[sheet_name].iter_rows(min_row=self.excel_min_row(sheet_name) + 1,
                                                   max_col=self.excel_max_col(sheet_name),
                                                   values_only=True)
        for value in all_values:
            case_list.append(case(*value))
        return case_list

    def write_cell(self, sheet_name, row, column, value, color='black'):
        """write cell value with color"""
        if isinstance(row, int) and isinstance(column, int):
            try:
                cell_obj = self.wb[sheet_name].cell(row, column)
                cell_obj.font = Font(color=self.RGBDict[color], bold=True)
                cell_obj.value = value
                self.wb.save(self.filename)
            except Exception as e:
                raise e
        else:
            raise TypeError('row and column must be type int')


do_excel = ParseExcel(DATA_PATH)

if __name__ == '__main__':
    do_excel = ParseExcel(DATA_PATH)
    print(do_excel.excel_max_col('sendMCode'))
    print(do_excel.excel_min_row('sendMCode'))
    print(do_excel.head('sendMCode'))
    print(do_excel.get_all_values('sendMCode'))
    do_excel.write_cell('sendMCode', 2, 8, 'test', color='red')
